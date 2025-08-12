import re
import io
from datetime import datetime

import streamlit as st
import pdfplumber
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font

# ---------------- Helpers: parsing response PDF ---------------- #

def extract_responses_from_bytes(pdf_bytes):
    """
    Parse response PDF bytes and return dict: qid -> set(['1','2',...])
    """
    responses = {}
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        text = "\n".join((page.extract_text() or "") for page in pdf.pages)

    # pattern finds Question ID ... Chosen Option ... (handles single/multiple)
    matches = re.findall(
        r"Question\s*ID\s*:\s*(\d+)[\s\S]*?Chosen\s*Option\s*:\s*([0-9,\|\s]+)",
        text,
        flags=re.IGNORECASE,
    )
    for qid, opttext in matches:
        opts = set(re.findall(r"\d+", opttext))
        responses[qid] = opts

    return responses


# ---------------- Helpers: parsing answer key (color-based) ---------------- #

def _span_color_to_rgb(color):
    """Return r,g,b ints from fitz span color (int OR tuple/list OR None)."""
    r = g = b = 0
    try:
        if color is None:
            return 0, 0, 0
        # int like 16711935 (0xRRGGBB)
        if isinstance(color, (int, float)):
            color_int = int(color)
            r = (color_int >> 16) & 255
            g = (color_int >> 8) & 255
            b = color_int & 255
            return r, g, b
        # tuple/list probably floats 0..1 or ints 0..255
        if isinstance(color, (list, tuple)):
            if all(0 <= c <= 1 for c in color[:3]):
                return int(color[0] * 255), int(color[1] * 255), int(color[2] * 255)
            else:
                return int(color[0]), int(color[1]), int(color[2])
    except Exception:
        pass
    return r, g, b


def extract_answerkey_with_colors_from_bytes(pdf_bytes):
    """
    Parse answer key PDF bytes using PyMuPDF to detect green-colored option numbers.
    Returns:
      - answerkey: dict qid -> set(['1','2',...])
      - ambiguous: set(qid) where 'candidate will get full marks' notes appear
    """
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    answerkey = {}
    ambiguous = set()
    current_qid = None

    for page in doc:
        blocks = page.get_text("dict").get("blocks", [])
        for b in blocks:
            for line in b.get("lines", []):
                # joined text of the line (to detect qid and ambiguity notes)
                line_text = "".join(s.get("text", "") for s in line.get("spans", [])).strip()

                # detect question id
                m_q = re.search(r"Question\s+Id\s*:\s*(\d+)", line_text, flags=re.IGNORECASE)
                if m_q:
                    current_qid = m_q.group(1)
                    answerkey.setdefault(current_qid, set())

                # detect ambiguous note (associate with current_qid)
                if current_qid and ("ambigu" in line_text.lower() or "candidate will get full marks" in line_text.lower()):
                    ambiguous.add(current_qid)

                # inspect individual spans for option numbers and colors
                for s in line.get("spans", []):
                    span_text = (s.get("text") or "").strip()
                    m_opt = re.match(r"^(\d+)\.\s*", span_text)
                    if m_opt and current_qid:
                        opt_num = m_opt.group(1)
                        color = s.get("color")
                        r, g, b = _span_color_to_rgb(color)
                        # detect green color (tune thresholds if necessary)
                        if (g > 120) and (g > r + 30) and (g > b + 30):
                            answerkey[current_qid].add(opt_num)

    doc.close()
    return answerkey, ambiguous


# ---------------- Scoring logic ---------------- #

def calculate_score(responses, answerkey, ambiguous):
    """
    Compare responses and answerkey.
    Return: correct_count, wrong_count, details_list
    details_list entries: (qid, chosen_set, correct_set, result_string)
    """
    correct = 0
    wrong = 0
    details = []

    for qid, chosen in responses.items():
        if qid not in answerkey:
            details.append((qid, chosen, set(), "no-key"))
            wrong += 1
            continue

        correct_set = answerkey[qid]
        chosen_set = set(chosen)

        if len(correct_set) == 1:
            if chosen_set == correct_set:
                correct += 1
                details.append((qid, chosen_set, correct_set, "correct"))
            else:
                wrong += 1
                details.append((qid, chosen_set, correct_set, "wrong"))
        else:
            # multiple-correct case
            if qid in ambiguous:
                # ambiguous -> any correct option chosen gives full marks
                if chosen_set & correct_set:
                    correct += 1
                    details.append((qid, chosen_set, correct_set, "correct (ambiguous)"))
                else:
                    wrong += 1
                    details.append((qid, chosen_set, correct_set, "wrong (ambiguous)"))
            else:
                # default: require exact match for multi-correct (you can change)
                if chosen_set == correct_set:
                    correct += 1
                    details.append((qid, chosen_set, correct_set, "correct (multi exact)"))
                else:
                    wrong += 1
                    details.append((qid, chosen_set, correct_set, "wrong (multi)"))

    return correct, wrong, details


# ---------------- Create Excel in-memory ---------------- #

def create_excel_bytes(details, correct, wrong, final_score):
    """
    Create an Excel workbook in memory and return its bytes (for download).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["Question ID", "Chosen Options", "Correct Options", "Result"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    for qid, chosen, correct_set, result in details:
        ws.append([
            qid,
            ", ".join(sorted(chosen)) if chosen else "",
            ", ".join(sorted(correct_set)) if correct_set else "",
            result
        ])

    # summary rows
    ws.append([])
    summary_start = ws.max_row + 1
    ws.append(["Summary"])
    ws.append(["Correct Answers", correct])
    ws.append(["Wrong Answers", wrong])
    ws.append(["Final Score (Correct/2)", final_score])

    # save to bytes
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


# ---------------- Streamlit UI ---------------- #

st.set_page_config(page_title="PDF Response Checker", layout="wide")
st.title("PDF Response Sheet & Answer Key Checker (Streamlit)")

st.markdown(
    """
Upload the student's **response sheet PDF** and the **answer key PDF** (where correct options are colored green).
- The app will detect green-colored options in the answer key as correct.
- It supports ambiguous questions marked like *'candidate will get full marks'* (in which case any correct option gives full marks).
- Final score = Correct / 2 (as requested).
"""
)

col1, col2 = st.columns(2)

with col1:
    resp_file = st.file_uploader("Upload Response Sheet PDF", type=["pdf"])

with col2:
    key_file = st.file_uploader("Upload Answer Key PDF", type=["pdf"])

process_clicked = st.button("Process PDFs")

if process_clicked:
    if not resp_file or not key_file:
        st.error("Please upload both PDFs.")
    else:
        try:
            # read uploaded files as bytes
            resp_bytes = resp_file.getvalue()
            key_bytes = key_file.getvalue()

            # parse
            with st.spinner("Parsing response sheet..."):
                responses = extract_responses_from_bytes(resp_bytes)

            with st.spinner("Parsing answer key (detecting green options)..."):
                answerkey, ambiguous = extract_answerkey_with_colors_from_bytes(key_bytes)

            correct, wrong, details = calculate_score(responses, answerkey, ambiguous)
            final_score = correct / 2.0

            st.success("Processing complete!")

            # show summary
            st.subheader("Summary")
            st.write(f"Correct answers: **{correct}**")
            st.write(f"Wrong answers: **{wrong}**")
            st.write(f"Final score (Correct / 2): **{final_score}**")

            # show ambiguous list if any
            if ambiguous:
                st.info(f"Ambiguous QIDs treated as 'any correct option gives full marks': {sorted(ambiguous)}")

            # show mismatches as a table
            st.subheader("All Results (sample / filterable)")
            # convert details to a list of dicts for nicer display
            rows = []
            for qid, chosen, corr, result in details:
                rows.append({
                    "Question ID": qid,
                    "Chosen": ", ".join(sorted(chosen)) if chosen else "",
                    "Correct": ", ".join(sorted(corr)) if corr else "",
                    "Result": result
                })
            st.dataframe(rows)

            # create excel bytes and provide download
            excel_bytes = create_excel_bytes(details, correct, wrong, final_score)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_name = f"results_{ts}.xlsx"
            st.download_button("Download results Excel", data=excel_bytes, file_name=default_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        except Exception as e:
            st.exception(e)
